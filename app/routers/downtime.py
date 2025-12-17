from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, extract, asc, desc
from datetime import datetime, date, timedelta, timezone
from typing import List, Optional
import io
from fastapi.responses import StreamingResponse

REASON_MAP = {
  "SETUP_DIE": "SETUP DIE",
  "REPAIR": "เครื่องขัดข้อง/alarm",
  "MAINTENANCE": "บำรุงรักษา",
  "MATERIAL_SHORTAGE": "รอวัตถุดิบ",
  "POWER_FAILURE": "ไฟฟ้าขัดข้อง",
  "QUALITY_CHECK": "ตรวจสอบคุณภาพ",
  "WAITING_APPROVAL": "รอการอนุมัติ",
  "OPERATOR_BREAK": "พักผ่อน",
  "OTHER_1": "อื่นๆ 1",
  "OTHER_2": "อื่นๆ 2",
}

def format_duration(seconds: Optional[int]) -> str:
    if seconds is None or seconds <= 0:
        return "00:00:00"
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"

from ..database import get_db
from ..models import DowntimeLog, DailySummary
from ..schemas import (
    DowntimeStartRequest, 
    DowntimeStopRequest, 
    DowntimeLogSchema,
    ActiveDowntimeResponse
)

router = APIRouter(prefix="/downtime", tags=["downtime"])

@router.post("/start", response_model=DowntimeLogSchema)
def start_downtime(request: DowntimeStartRequest, db: Session = Depends(get_db)):
    """เริ่มบันทึก downtime ใหม่"""
    # ตรวจสอบว่ามี downtime ที่ active อยู่หรือไม่
    active_downtime = db.query(DowntimeLog).filter(
        DowntimeLog.is_active == True
    ).first()
    
    if active_downtime:
        raise HTTPException(status_code=400, detail="มี downtime ที่กำลัง active อยู่แล้ว กรุณาหยุดก่อน")
    
    # สร้าง downtime log ใหม่
    new_downtime = DowntimeLog(
        downtime_reason=request.downtime_reason,
        date=date.today(),
        is_active=True
    )
    
    db.add(new_downtime)
    db.commit()
    db.refresh(new_downtime)
    
    return new_downtime

@router.post("/stop", response_model=DowntimeLogSchema)
def stop_downtime(db: Session = Depends(get_db)):
    """หยุดบันทึก downtime ปัจจุบัน"""
    # หา downtime ที่ active อยู่
    active_downtime = db.query(DowntimeLog).filter(
        DowntimeLog.is_active == True
    ).first()
    
    if not active_downtime:
        raise HTTPException(status_code=404, detail="ไม่พบ downtime ที่กำลัง active อยู่")
    
    # อัพเดท end_time และคำนวณ duration
    active_downtime.end_time = datetime.now(timezone.utc)
    duration = (active_downtime.end_time - active_downtime.start_time).total_seconds()
    active_downtime.duration_sec = int(duration)
    active_downtime.is_active = False
    
    # อัพเดท daily summary
    summary = db.query(DailySummary).filter(
        DailySummary.date == active_downtime.date
    ).first()
    
    if summary:
        summary.total_downtime_sec += active_downtime.duration_sec
    else:
        summary = DailySummary(
            date=active_downtime.date,
            total_cycles=0,
            total_runtime_sec=0,
            total_downtime_sec=active_downtime.duration_sec,
            availability=0.0
        )
        db.add(summary)
    
    SHIFT_SECONDS = 27000.0
    summary.availability = round(min(100.0, summary.total_runtime_sec / SHIFT_SECONDS * 100), 2)
    
    db.commit()
    db.refresh(active_downtime)
    
    return active_downtime

@router.get("/active", response_model=ActiveDowntimeResponse)
def get_active_downtime(db: Session = Depends(get_db)):
    """ดึงข้อมูล downtime ที่กำลัง active อยู่"""
    active_downtime = db.query(DowntimeLog).filter(
        DowntimeLog.is_active == True
    ).first()
    
    return {
        "is_active": active_downtime is not None,
        "current_downtime": active_downtime
    }

@router.get("/summary/today")
def get_today_downtime_summary(db: Session = Depends(get_db)):
    """ดึงข้อมูลสรุป downtime แต่ละประเภทสำหรับวันนี้"""
    today = date.today()
    
    # Query downtime logs for today grouped by reason
    from sqlalchemy import func
    results = db.query(
        DowntimeLog.downtime_reason,
        func.sum(DowntimeLog.duration_sec).label("total_duration_sec")
    ).filter(
        and_(
            DowntimeLog.date == today,
            DowntimeLog.is_active == False,
            DowntimeLog.duration_sec.isnot(None)
        )
    ).group_by(DowntimeLog.downtime_reason).all()
    
    # Convert to dictionary
    summary = {}
    for result in results:
        summary[result.downtime_reason] = result.total_duration_sec or 0
    
    # Ensure all reasons are present (even if 0)
    for reason_id in REASON_MAP.keys():
        if reason_id not in summary:
            summary[reason_id] = 0
    
    return summary

@router.get("/history", response_model=List[DowntimeLogSchema])
def get_downtime_history(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    """ดึงประวัติ downtime"""
    query = db.query(DowntimeLog)
    
    if start_date:
        query = query.filter(DowntimeLog.date >= start_date)
    if end_date:
        query = query.filter(DowntimeLog.date <= end_date)
    
    downtimes = query.order_by(DowntimeLog.start_time.desc()).limit(limit).all()
    return downtimes

@router.get("/top-today", response_model=List[DowntimeLogSchema])
def get_top_downtime_today(db: Session = Depends(get_db)):
    """ดึง Top 10 Downtime วันนี้ เรียงตามระยะเวลายาวนานที่สุด"""
    today = date.today()
    
    top_downtimes = db.query(DowntimeLog).filter(
        and_(
            DowntimeLog.date == today,
            DowntimeLog.is_active == False,
            DowntimeLog.duration_sec.isnot(None)
        )
    ).order_by(
        desc(DowntimeLog.duration_sec),
        desc(DowntimeLog.start_time)
    ).limit(10).all()
    
    return top_downtimes

@router.get("/export")
def export_downtime_report(
    report_type: str,
    year: int,
    month: Optional[int] = None,
    day: Optional[int] = None,
    db: Session = Depends(get_db)
):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    wb = openpyxl.Workbook()
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    if report_type == "daily":
        if not month or not day:
            raise HTTPException(status_code=400, detail="Daily report requires month and day")
        
        target_date = date(year, month, day)
        filename = f"Daily_Report_{target_date.strftime('%Y-%m-%d')}.xlsx"
        
        # Sheet 1: Downtime_Log
        ws1 = wb.active
        ws1.title = "Downtime_Log"
        headers1 = ["ลำดับ (No.)", "เวลาเริ่ม (Start Time)", "เวลาหยุด (End Time)", "สาเหตุ (Reason)", "ระยะเวลา (Duration)"]
        for col_num, header_text in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col_num, value=header_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        downtimes = db.query(DowntimeLog).filter(
            DowntimeLog.date == target_date
        ).order_by(DowntimeLog.start_time.asc()).all()
        
        for idx, downtime in enumerate(downtimes, start=2):
            ws1.cell(row=idx, column=1, value=idx - 1)
            ws1.cell(row=idx, column=2, value=downtime.start_time.strftime("%H:%M:%S"))
            end_time_str = downtime.end_time.strftime("%H:%M:%S") if downtime.end_time else "-"
            ws1.cell(row=idx, column=3, value=end_time_str)
            reason_label = REASON_MAP.get(downtime.downtime_reason, downtime.downtime_reason)
            ws1.cell(row=idx, column=4, value=reason_label)
            duration_min = round(downtime.duration_sec / 60, 1) if downtime.duration_sec else "-"
            ws1.cell(row=idx, column=5, value=duration_min)
        
        # Sheet 2: Machine_Summary
        ws2 = wb.create_sheet("Machine_Summary")
        headers2 = ["วันที่", "Total Cycles (เพิ่ม)", "Total Runtime", "Total Downtime", "Availability (%)"]
        for col_num, header_text in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_num, value=header_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        summary = db.query(DailySummary).filter(DailySummary.date == target_date).first()
        cycles = summary.total_cycles if summary else 0
        runtime_sec = summary.total_runtime_sec if summary else 0
        downtime_sec = summary.total_downtime_sec if summary else 0
        total_time_sec = runtime_sec + downtime_sec
        availability_pct = round((runtime_sec / total_time_sec * 100), 2) if total_time_sec > 0 else 100.0
        
        row = 2
        ws2.cell(row=row, column=1, value=target_date.strftime("%Y-%m-%d"))
        ws2.cell(row=row, column=2, value=cycles)
        ws2.cell(row=row, column=3, value=format_duration(runtime_sec))
        ws2.cell(row=row, column=4, value=format_duration(downtime_sec))
        ws2.cell(row=row, column=5, value=f"{availability_pct:.2f}")
    
    elif report_type == "monthly":
        if not month:
            raise HTTPException(status_code=400, detail="Monthly report requires month")
        
        filename = f"Monthly_Report_{year}-{month:02d}.xlsx"
        
        # Sheet 1: Daily_Performance
        ws1 = wb.active
        ws1.title = "Daily_Performance"
        headers1 = ["วันที่ (Date)", "Total Cycles (จำนวนผลิต)", "Total Runtime (เวลาเดินเครื่อง)", "Total Downtime (เวลาหยุดเครื่อง)", "Availability % (ประสิทธิภาพ)"]
        for col_num, header_text in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col_num, value=header_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        total_downtime_month_sec = 0
        
        for day_num in range(1, 32):
            try:
                day_date = date(year, month, day_num)
            except ValueError:
                continue
            
            summary = db.query(DailySummary).filter(DailySummary.date == day_date).first()
            cycles = summary.total_cycles if summary else 0
            runtime_sec = summary.total_runtime_sec if summary else 0
            downtime_sec = summary.total_downtime_sec if summary else 0
            
            total_time_sec = runtime_sec + downtime_sec
            availability_pct = round((runtime_sec / total_time_sec * 100), 2) if total_time_sec > 0 else 100.0
            
            row = day_num + 1
            ws1.cell(row=row, column=1, value=day_num)
            ws1.cell(row=row, column=2, value=cycles)
            ws1.cell(row=row, column=3, value=format_duration(runtime_sec))
            ws1.cell(row=row, column=4, value=format_duration(downtime_sec))
            ws1.cell(row=row, column=5, value=f"{availability_pct:.2f}")
            
            total_downtime_month_sec += downtime_sec
        
        # Sheet 2: Top_Downtime_Reasons
        ws2 = wb.create_sheet("Top_Downtime_Reasons")
        headers2 = ["สาเหตุ (Reason)", "จำนวนครั้ง (Frequency)", "เวลารวม (Total Duration)", "% ของเวลาที่เสียไป"]
        for col_num, header_text in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_num, value=header_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        reasons_data = db.query(
            DowntimeLog.downtime_reason,
            func.count(DowntimeLog.id).label("frequency"),
            func.sum(DowntimeLog.duration_sec).label("total_duration_sec")
        ).filter(
            and_(
                func.extract("year", DowntimeLog.date) == year,
                func.extract("month", DowntimeLog.date) == month,
                DowntimeLog.is_active == False,
                DowntimeLog.duration_sec.isnot(None)
            )
        ).group_by(DowntimeLog.downtime_reason).order_by(desc("total_duration_sec")).all()
        
        for idx, row_data in enumerate(reasons_data, start=2):
            reason_label = REASON_MAP.get(row_data.downtime_reason, row_data.downtime_reason)
            frequency = row_data.frequency
            total_duration_min = round(row_data.total_duration_sec / 60, 1) if row_data.total_duration_sec else 0
            pct = round((row_data.total_duration_sec / total_downtime_month_sec * 100), 2) if total_downtime_month_sec > 0 else 0.0
            
            ws2.cell(row=idx, column=1, value=reason_label)
            ws2.cell(row=idx, column=2, value=frequency)
            ws2.cell(row=idx, column=3, value=total_duration_min)
            ws2.cell(row=idx, column=4, value=f"{pct:.2f}")
    
    elif report_type == "yearly":
        filename = f"Yearly_Report_{year}.xlsx"
        
        month_names = ["ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.", "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]
        
        # Sheet 1: Monthly_Performance
        ws1 = wb.active
        ws1.title = "Monthly_Performance"
        headers1 = ["เดือน (Month)", "Total Cycles (ยอดผลิตรวม)", "Total Runtime (ชั่วโมงทำงานรวม)", "Total Downtime (ชั่วโมงหยุดเครื่องรวม)", "Avg Availability % (ค่าเฉลี่ยประสิทธิภาพ)"]
        for col_num, header_text in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col_num, value=header_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        monthly_summaries = db.query(
            func.extract("month", DailySummary.date).label("month"),
            func.coalesce(func.sum(DailySummary.total_cycles), 0).label("total_cycles"),
            func.coalesce(func.sum(DailySummary.total_runtime_sec), 0).label("total_runtime_sec"),
            func.coalesce(func.sum(DailySummary.total_downtime_sec), 0).label("total_downtime_sec")
        ).filter(
            func.extract("year", DailySummary.date) == year
        ).group_by(func.extract("month", DailySummary.date)).order_by("month").all()
        
        total_year_downtime_sec = sum((ms.total_downtime_sec or 0) for ms in monthly_summaries)
        
        for month_idx in range(12):
            month_num = month_idx + 1
            month_row = next((ms for ms in monthly_summaries if int(ms.month) == month_num), None)
            cycles = month_row.total_cycles if month_row else 0
            runtime_sec = month_row.total_runtime_sec if month_row else 0
            downtime_sec = month_row.total_downtime_sec if month_row else 0
            total_time_sec = runtime_sec + downtime_sec
            avg_availability = round((runtime_sec / total_time_sec * 100), 2) if total_time_sec > 0 else 100.0
            
            row = month_idx + 2
            ws1.cell(row=row, column=1, value=month_names[month_idx])
            ws1.cell(row=row, column=2, value=cycles)
            ws1.cell(row=row, column=3, value=format_duration(runtime_sec))
            ws1.cell(row=row, column=4, value=format_duration(downtime_sec))
            ws1.cell(row=row, column=5, value=f"{avg_availability:.2f}")
        
        # Sheet 2: Top_Downtime_Reasons_Yearly
        ws2 = wb.create_sheet("Top_Downtime_Reasons_Yearly")
        headers2 = ["สาเหตุ (Reason)", "จำนวนครั้ง (Frequency)", "เวลารวม (Total Hours)", "% Impact (เทียบกับเวลาหยุดทั้งหมดของปี)"]
        for col_num, header_text in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_num, value=header_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        yearly_reasons = db.query(
            DowntimeLog.downtime_reason,
            func.count(DowntimeLog.id).label("frequency"),
            func.sum(DowntimeLog.duration_sec).label("total_duration_sec")
        ).filter(
            and_(
                func.extract("year", DowntimeLog.date) == year,
                DowntimeLog.is_active == False,
                DowntimeLog.duration_sec.isnot(None)
            )
        ).group_by(DowntimeLog.downtime_reason).order_by(desc("total_duration_sec")).all()
        
        total_year_downtime_from_reasons = sum((r.total_duration_sec or 0) for r in yearly_reasons)
        
        for idx, row_data in enumerate(yearly_reasons, start=2):
            reason_label = REASON_MAP.get(row_data.downtime_reason, row_data.downtime_reason)
            frequency = row_data.frequency
            total_hours = round(row_data.total_duration_sec / 3600, 1) if row_data.total_duration_sec else 0
            pct = round((row_data.total_duration_sec / total_year_downtime_from_reasons * 100), 2) if total_year_downtime_from_reasons > 0 else 0.0
            
            ws2.cell(row=idx, column=1, value=reason_label)
            ws2.cell(row=idx, column=2, value=frequency)
            ws2.cell(row=idx, column=3, value=total_hours)
            ws2.cell(row=idx, column=4, value=f"{pct:.2f}")
    
    else:
        raise HTTPException(status_code=400, detail="Invalid report_type. Use 'daily', 'monthly', or 'yearly'")
    
    # Auto adjust column widths for all sheets
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value or "")) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
